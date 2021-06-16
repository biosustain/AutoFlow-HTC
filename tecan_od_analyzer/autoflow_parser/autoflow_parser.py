# -*- coding: utf-8 -*-

"""Main module."""
import pandas as pd

from string import ascii_uppercase
from openpyxl import load_workbook
from datetime import datetime, timedelta
from os.path import join
from collections import OrderedDict


def process_logfile(path: str):
    """Check logfile and return list of entries."""
    header_columns = [
        "Sampling Time",
        "Measurement Type",
        "Bioshaker Processed",
        "Sampling ID",
        "Quadrant Used in Reading Plate",
        "Dilution Factor",
    ]

    log_df = pd.read_csv("Logfile.txt", sep=";")

    # Check if logfile has required column headers
    missing = set(header_columns) - set(log_df.columns)
    if missing != set():
        return (
            -1,
            f"Missing required column(s) in Logfile.txt: {', '.join(missing)}",
        )
    log_df = log_df[header_columns]
    entries = (
        log_df.astype(str)
        .apply(lambda row: ";".join(row.values), axis=1)
        .to_list()
    )
    if "Experiment Finished" in entries[-1]:
        entries = entries[0:-1]
    return (0, entries)


def check_results_files(entries: list, files: set, max_timeflex: int):
    """A function to check results files wrt logfile"""
    entry_to_file = {}

    # Filenames can have multiple formats based on date/time used.
    naming_formats = [
        "%Y_%m_%d_%H_%M_%S",
        "%Y_%m_%d_%I_%M_%S",
        "%Y_%m_%d_%H_%M",
        "%Y_%m_%d_%I_%M",
    ]

    tried_timeflex = [max_timeflex]
    timeflex = max_timeflex
    org_files = files
    while True:
        missing_files, multiple_mathces = False, False
        files = org_files
        for entry in entries:
            (timestamp, measurement) = entry.split(";")[0:2]
            timestamp = datetime.strptime(timestamp, "%d-%m-%Y %H:%M:%S")

            # Filenames can be multiple seconds off from what's logged
            # in logfile.
            # Adjust expected filename based on allowed flexibility.
            filenames = set(
                [
                    "_".join([measurement, time.strftime(nf)]) + ".xlsx"
                    for time in [
                        timestamp + timedelta(seconds=i)
                        for i in range(timeflex + 1)
                    ]
                    for nf in naming_formats
                ]
            )

            matches = filenames & files
            if len(matches) == 0:
                missing_files = True
                entry_to_file[entry] = ""
            elif len(matches) == 1:
                files = files - matches
                entry_to_file[entry] = matches.pop()
            else:
                multiple_mathces = True
                # files = files - matches
                entry_to_file[entry] = matches

        if not (missing_files ^ multiple_mathces):
            # nothing to do if they are both True
            # no need to do anything if they are both false
            break
        else:
            # try using a smaller timeflex if there are multiple mathces and
            # larger if missing files.
            if multiple_mathces:
                timeflex = min(max(timeflex - 5, 0), min(tried_timeflex))
                if timeflex in tried_timeflex:
                    break
                else:
                    tried_timeflex.append(timeflex)
            elif missing_files:
                timeflex = min(timeflex + 1, max_timeflex)
                if timeflex in tried_timeflex:
                    break
                else:
                    tried_timeflex.append(timeflex)

    return (timeflex, missing_files, multiple_mathces, entry_to_file, files)


def parse_tecan_files(entry: str, file: str, path: str):
    """Method to parse tecan files and return pandas dataframe"""
    # pick up details from logfile entry
    entry = entry.split(";")
    details = OrderedDict(
        [
            ("Sampling date", entry[0].split()[0]),
            ("Sampling time", entry[0].split()[1]),
            ("Measurement type", entry[1]),
            ("Bioshaker processed", entry[2]),
            ("Sampling ID", entry[3]),
            ("Quadrant", entry[4]),
            ("Dilution Factor", entry[5]),
        ]
    )

    # read in the workbook and pick up more details
    wb = load_workbook(join(path, file), read_only=True, data_only=True)
    ws = wb["Sheet2"]  # seems to be the standard
    lambdas = {
        "Application": lambda row: {
            "Application": row[0].value.split(": ")[1]
        },
        "Device": lambda row: {
            "Device": row[0].value.split(": ")[1],
            "Serial number": row[4].value.split(": ")[1],
        },
        "Firmware": lambda row: {"Firmware": row[0].value.split(": ")[1]},
        "System": lambda row: {"System": row[4].value},
        "User": lambda row: {"User": row[4].value},
        "Plate": lambda row: {"Plate": row[4].value}
        if row[0].value == "Plate"
        else {},
        "Shaking (Linear) Duration": lambda row: {
            f"Shaking (Linear) Duration ({row[5].value})": row[4].value
        },
        "Shaking (Linear) Amplitude": lambda row: {
            f"Shaking (Linear) Amplitude ({row[5].value})": row[4].value
        },
        "Mode": lambda row: {"Mode": row[4].value},
        "Wavelength": lambda row: {
            f"Wavelength ({row[5].value})": row[4].value
        },
        "Bandwidth": lambda row: {f"Bandwidth ({row[5].value})": row[4].value},
        "Number of Flashes": lambda row: {"Number of Flashes": row[4].value},
        "Settle Time": lambda row: {
            f"Settle Time ({row[5].value})": row[4].value
        },
        "Start Time": lambda row: {"Start Time": row[1].value},
        "Temperature": lambda row: {
            "Temperature": row[1].value.split(": ")[1]
        },
        "<>": lambda row: {"<>": (row[0].row, row[0].column)},
        "End Time": lambda row: {"End Time": row[1].value},
    }
    for row in ws.iter_rows():
        test_str = str(row[0].value) + str(row[1].value)
        for key, value in lambdas.items():
            if test_str.startswith(key):
                details.update(value(row))
                # del lambdas[key]  # can't do this to eliminate found keys.
                continue

    # read the measured values table into dataframe
    df = pd.melt(
        pd.read_excel(
            join(path, file), skiprows=details["<>"][0] - 1, skipfooter=1
        )
        .dropna()
        .rename(columns={"<>": "Row"}),
        id_vars="Row",
        var_name="Column",
        value_name="Measurement",
    )

    del details["<>"]
    # add details to dataframe
    for key, value in reversed(details.items()):
        df.insert(0, key, value)

    # insert columns to match sample row/column from 96-well plate
    df.insert(0, "Sample row", "")
    df["Sample row"] = df.Row.apply(
        lambda row: ascii_uppercase[ascii_uppercase.index(row) // 2]
    )
    df.insert(1, "Sample column", "")
    df["Sample column"] = df.Column.apply(lambda column: (column + 1) // 2)

    # TODO: This sample name is only temporary and needs to match the value
    #  from a "layout" file once it's ready.
    df.insert(0, "Sample ID", "")
    df["Sample ID"] = df.apply(
        lambda row: f"BS{row['Bioshaker processed']}_{row['Sample row']}{row['Sample column']}",  # noqa E501
        axis=1,
    )

    return df


def merge_results(entry_to_file: dict, path: str):
    """Combine measurement dataframes into one"""
    df = pd.DataFrame()
    for entry, file in entry_to_file.items():
        df = pd.concat(
            [df, parse_tecan_files(entry, file, path)],
            ignore_index=True,
            sort=True,
        )
    return df


def make_channel_dfs(merged_df: pd.DataFrame):
    """Make dataframes with only elapsed time (h) and samples for plotting"""
    # only take what you need
    tmp_df = merged_df[
        [
            "Sample ID",
            "Measurement type",
            "Start Time",
            "Measurement",
            "Bioshaker processed",
        ]
    ].copy()

    # When the detector gets oversaturated for a measurement, it reports "OVER"
    # Jerome and I decided to convert that an artificially high number
    # (1000000) in order to keep all values numeric.
    tmp_df.Measurement.replace("OVER", 1000000, inplace=True)

    # Convert "Start Time" to "Elapsed time (h)"
    tmp_df["Start Time"] = pd.to_datetime(
        tmp_df["Start Time"], format="%d-%m-%Y %H:%M:%S"
    )
    start_time = tmp_df["Start Time"].min()
    tmp_df["Elapsed time (h)"] = tmp_df["Start Time"].apply(
        lambda time: (time - start_time).total_seconds() / 3600
    )

    # create regular tables of time vs sample measurement for each channel
    channel_dfs = {}
    for bioshaker in tmp_df["Bioshaker processed"].unique():
        for channel in tmp_df["Measurement type"].unique():
            channel_dfs[f"BS{bioshaker}_{channel}"] = (
                tmp_df[
                    (tmp_df["Measurement type"] == channel)
                    & (tmp_df["Bioshaker processed"] == bioshaker)
                ]
                .pivot_table(
                    values="Measurement",
                    index="Elapsed time (h)",
                    columns="Sample ID",
                )
                .sort_index()
            )

    return channel_dfs
